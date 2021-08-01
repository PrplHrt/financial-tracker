# class Customer():
#     def __init__(self, n: str, i: int) -> None:
#         self.name = n
#         self.ID = i

# class Item():
#     def __init__(self, d:str, p: float) -> None:
#         self.description = d
#         self.price = p

#     def get_description(self):
#         return self.description

#     def get_price(self):
#         return self.price

# class Order():
#     def __init__(self, cust: Customer) -> None:
#         self.customer = cust
#         self.items = []
#         self.shipment = 0.
#         self.total = 0.

#     def add_item(self, item):
#         self.items.append(item)
#         self.total += item.price

#     def set_shipment(self, ship: float):
#         self.shipment = ship

#     def get_order(self):
#         # Function to return readable format
#         pass

from openpyxl import Workbook

# Create a workbook
wb = Workbook()

# Get the worksheet in the workbook
ws = wb.active

# Changing the worksheet title
ws.title = "8_1_2021"

# Changing tab color
ws.sheet_properties.tabColor = "C8A2C8"

for i in range(3):
    name, item, price = input().split()
    n =ws.cell(row = i+1, column=1)
    n.value = name
    t = ws.cell(row = i+1, column=2)
    t.value = item
    p = ws.cell(row = i+1, column=3) 
    p.value = price


# Saving the workbook
wb.save("test.xlsx")
