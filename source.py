from datetime import date
from openpyxl import Workbook, load_workbook
from order import Order
from insertOrder import OrderFrame
import pandas as pd
# Program that will prep sheet: set title as date, set headings

# Orders
orders = []

def set_worksheet(ws):
    # Setting the worksheet title
    ws['A1'] = "Customer Name"
    ws['B1'] = "Item Description"
    ws['C1'] = "Amount"
    ws['D1'] = "Total"
    ws['E1'] = "Eligible for SF"
    ws['F1'] = "NET TOTAL"
    ws['G1'] = "Remarks"
    ws['H1'] = "Payment Method"

def load_worksheet(ws):
    global orders
    i = 2
    while ws.cell(row = i, column = 1).value != "":
        order = Order(ws.cell(row = i, column = 1).value)
        i += 1
        while  ws.cell(row = i, column = 4).value == "" and ws.cell(row = i, column = 2).value != "" and ws.cell(row = i, column = 3).value != "":
            order.insert_item(ws.cell(row = i, column = 2).value, ws.cell(row = i, column = 3).value)
            i += 1
        if ws.cell(row = i, column = 2).value != "" and ws.cell(row = i, column = 3).value != "":
            order.insert_item(ws.cell(row = i, column = 2).value, ws.cell(row = i, column = 3).value)
        order.eligible(order.total<150)
        i += 3
        orders.append(order)

def save_worksheet(ws):
    i = 2
    for order in orders:
        ws.cell(row = i, column = 1).value = order.cust_name
        i += 1
        for item in order.items:
            ws.cell(row = i, column = 2).value = item["desc"]
            ws.cell(row = i, column = 3).value = item["price"]
            i += 1
        ws.cell(row = i-1, column = 4).value = order.total
        ws.cell(row = i-1, column = 5).value = order.eligible_sf
        i += 2
    for order in orders:
        print(order.cust_name)
        print(order.items)
        print(order.total)
        print()

def make_order():
    print("Enter customer name:")
    name = str(input())
    order = Order(name)
    print("How many items to be added:")
    num = int(input())
    for i in range(num):
        of = OrderFrame(name)
        order.insert_item(str(of.item_name.get()), float(of.item_price.get()))
    orders.append(order)


if __name__ == '__main__':
    try:
        # Load already existing database
        wb = load_workbook('orders.xlsx')
        if str(date.today()) in wb.sheetnames:
            # If today's sheet exists, load it
            ws = wb[str(date.today())]
            load_worksheet(ws)
        else:
            # If today's sheet doesn't exist, make one
            ws = wb.create_sheet(str(date.today()))
            set_worksheet(ws)
    except FileNotFoundError:
        # Create new if database isn't there
        wb = Workbook()
        ws = wb.active
        ws.title = str(date.today())
        set_worksheet(ws)

    make_order()

    save_worksheet(ws)
    wb.save('orders.xlsx')